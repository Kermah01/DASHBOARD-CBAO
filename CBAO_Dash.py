import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl
from openpyxl import load_workbook
import numpy as np
from streamlit_extras.metric_cards import style_metric_cards # beautify metric card with css
import plotly.graph_objects as go
import nltk
from nltk import word_tokenize, ngrams
from collections import Counter
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import re
from unidecode import unidecode
import streamlit_gsheets
from streamlit_gsheets import GSheetsConnection
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)
def dashboard_users(user_zone):
    def base_locale():
        url= "https://docs.google.com/spreadsheets/d/1yy4k-xZCHWD4fxSkem8hQJ-nut3LyTh3J6-or7_8mRc/edit?usp=sharing"
        conn = st.connection("gsheets", type=GSheetsConnection)
        return conn.read(spreadsheet=url)
    
    if user_zone=='cbao-qualité':
        df=base_locale()
    else:
        df=base_locale()[base_locale()['Zone']==user_zone]
    


    def read_excel_file(file):
        data = load_workbook(file)
        datas = data.active
        donnees = []
        for ligne in datas.iter_rows(values_only=True):
            donnees.append(list(ligne))
        en_tetes = donnees[0]
        donnees = donnees[1:]
        new_df = pd.DataFrame(donnees, columns=en_tetes)
        return new_df

    st.sidebar.image('CBAO_GAWB_logo.jpg', use_column_width='always')
    uploaded_file = st.sidebar.file_uploader("Télécharger un fichier Excel", type=["xlsx"])

    # Vérifier si un fichier a été téléchargé
    if uploaded_file is not None:
        # Utiliser la fonction pour lire le fichier Excel
        df_uploaded = read_excel_file(uploaded_file)




    pec_agence=df.groupby('Agence')['Note de la prise en charge'].mean()
    order_of_months = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    dic_month={1:"Janvier",2:"Février",3:"Mars",4:"Avril",5:"Mai",6:"Juin",7:"Juillet",8:"Août",9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre"}
    order_of_days = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']

    
    def transf_df(df):
        #df['Horodateur'] = df['Horodateur'].astype(str).replace(':00', '')

        df['Horodateur']= pd.to_datetime(df['Horodateur'],format="%d/%m/%Y %H:%M", errors='coerce')
        df["Mois"] = df["Horodateur"].dt.month
        df["Jour"] = df["Horodateur"].dt.day_of_week
        df["heure"]=df["Horodateur"].dt.hour
        df["Année"]=df["Horodateur"].dt.year
        df["Mois"]=df["Mois"].map(dic_month)
        df["Mois*"] = pd.Categorical(df["Mois"], categories=order_of_months, ordered=True)
        df["Jour"]=df["Jour"].map({0:"Lundi",1:"Mardi",2:"Mercredi",3:"Jeudi",4:"Vendredi",5:"Samedi",6:"Dimanche"})
        df["Jour*"] = pd.Categorical(df["Jour"], categories=order_of_days, ordered=True)
        order_of_months_year = []
        start_year = df['Année'].min()
        end_year = df['Année'].max()
        for year in range(start_year, end_year + 1):
            for month in order_of_months:
                order_of_months_year.append(f"{month} {year}")

        df["Mois de l'année"] = df['Mois*'].astype("str") + ' ' + df['Année'].astype("str")
        df["Mois de l'année"] = pd.Categorical(df["Mois de l'année"], categories=order_of_months_year, ordered=True)
        
        return df
    
    df=transf_df(df)

    def filter_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """
        Adds a UI on top of a dataframe to let viewers filter columns

        Args:
            df (pd.DataFrame): Original dataframe

        Returns:
            pd.DataFrame: Filtered dataframe
        """
        modify = st.checkbox("AJOUTEZ UN FILTRE")
        

        if not modify:
            return df

        df = df.copy()

        # Try to convert datetimes into a standard format (datetime, no timezone)
        for col in df.columns:
            if is_object_dtype(df[col]):
                try:
                    df[col] = pd.to_datetime(df[col])
                except Exception:
                    pass

            if is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

        modification_container = st.container()

        with modification_container:
            to_filter_columns = st.multiselect("Choisissez les variables que vous souhaitez utiliser comme filtre", df.columns)
            for column in to_filter_columns:
                left, right = st.columns((1, 20))
                # Treat columns with < 10 unique values as categorical
                int_columns = df.select_dtypes(include="int").columns
                float_columns = df.select_dtypes(include="float").columns

                if is_numeric_dtype(df[column]) :
                    _min = int(df[column].min())
                    _max = int(df[column].max())
                    user_num_input = right.slider(
                        f"Valeurs de {column}",
                        min_value=_min,
                        max_value=_max,
                        value=(_min, _max),
                    )
                    df = df[df[column].between(*user_num_input)]
                elif is_datetime64_any_dtype(df[column]):
                    user_date_input = right.date_input(
                        f"Valeur de {column}",
                        value=(
                            df[column].min(),
                            df[column].max(),
                        ),
                    )
                    if len(user_date_input) == 2:
                        user_date_input = tuple(map(pd.to_datetime, user_date_input))
                        start_date, end_date = user_date_input
                        df = df.loc[df[column].between(start_date, end_date)]
                elif is_categorical_dtype(df[column]) or df[column].unique().shape[0]<100:
                    arr=df[column].unique()
                    user_cat_input = right.multiselect(
                        f"Valueur de {column}",
                        arr
                        ,
                        default=list(arr),
                    )
                    df = df[df[column].isin(user_cat_input)]
                else:
                    user_text_input = right.text_input(
                        f"Substring or regex in {column}",
                    )
                    if user_text_input:
                        df = df[df[column].astype(str).str.contains(user_text_input)]

        return df


    page_bg_img = f"""
    <style>
    [data-testid="stAppViewContainer"] > .main {{
    background-image: url(https://static.vecteezy.com/ti/vecteur-libre/p1/20530242-abstrait-arriere-plan-vague-doubler-violet-vague-colore-lignes-neon-lumiere-abstrait-fond-d-ecran-numerique-abstrait-3d-technologie-vague-effet-embrase-lignes-vague-arriere-plan-vectoriel.jpg);
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    background-attachment: no-fixed;
    height: 100vh;
    margin: 0;
    display: flex;


    }}
    [data-testid="stSidebar"] {{
        background-color: #000 !important;  /* Fond noir */
        border: 2px solid #f7a900 !important;  /* Bordure rouge */
        border-radius: 10px;  /* Coins arrondis */
        margin-top: 0 px;  /* Ajuster la position vers le haut */
        position: relative;
        z-index: 1;  /* S'assurer que la barre latérale est au-dessus du contenu */
        padding: 10px;
    }}

    [data-testid="stHeader"] {{
    background: rgba(0, 0, 0, 0);
    color: white;
    }}

    [data-testid="stToolbar"] {{
    right: 2rem;
    }}
    </style>
    """

    st.markdown(
        """
        <style>
            body {
                color: white;
            }
        </style>
        """,
        unsafe_allow_html=True
    )

    
    colors = px.colors.sequential.Rainbow_r
    colors.extend(px.colors.sequential.Agsunset)
    colors.extend(px.colors.sequential.Aggrnyl)
    # Section des graphiques sommaires






    st.write("\n")
    st.write("\n")

    st.markdown(page_bg_img, unsafe_allow_html=True)

    dec_temp1=dec_temp2=dec_temp=df["Mois*"].unique().sort_values(ascending=False)[1]


    acc_selected=pec_selected=''

    st.sidebar.subheader("PARAMETRES DES VARIATIONS DES KPI")
    st.title("KPI annuels")
    year=st.selectbox("Sélectionner l'année sur laquelle vous souhaitez réaliser votre analyse", np.sort(df['Année'].unique())[::-1],index=0)
    df_transitoire=df.copy()
    df=df[df['Année']==year]
    nbre_quest=st.sidebar.checkbox('Nombre total des questionnaires')
    if nbre_quest:
        dec_temp=st.sidebar.select_slider(":grey[***Choisissez le décalage temporel (en mois)***]", options=df["Mois*"].unique().sort_values(ascending=True), value=df["Mois*"].unique().sort_values(ascending=False)[1])

    moy_acc=st.sidebar.checkbox("Note moyenne de l'accueil") 
    if moy_acc:
        acc_selected=st.sidebar.radio(":grey[Modifiez le décalage temporel]", ['Modifier le décalage temporel', 'Modifier la norme'])
        
        if acc_selected=='Modifier le décalage temporel':
            dec_temp1=st.sidebar.select_slider(":grey[***Choisissez le décalage temporel (en mois)*** ]", options=df["Mois*"].unique().sort_values(ascending=True), value=df["Mois*"].unique().sort_values(ascending=False)[1])

        else: 
            new_norm=st.sidebar.number_input(":grey[***entrez la nouvelle norme***]", min_value=1, max_value=5, value=4)

    moy_pec=st.sidebar.checkbox("Note moyenne de la prise en charge")
    if moy_pec:
        pec_selected=st.sidebar.radio(":grey[Modifiez le décalage temporel]", ['Modifier le décalage temporel ', 'Modifier la norme '])
        if pec_selected=='Modifier le décalage temporel ':
            dec_temp2=st.sidebar.select_slider(":grey[***Choisissez le décalage temporel (en mois)***  ]", options=df["Mois*"].unique().sort_values(ascending=True), value=df["Mois*"].unique().sort_values(ascending=False)[1])
        else: 
            new_norm2=st.sidebar.number_input(":grey[***Entrez la nouvelle norme***  ]", min_value=1, max_value=5, value=4)

    
    st.header(f"***KPI sur l'année {year}***",divider ="rainbow" )
    with st.container():    
        col1, col2, col3 = st.columns(3)
        col1.metric(f"Total des questionnaires soumis en {year}", df.shape[0],f"{df.shape[0]-df[df['Mois*']<=dec_temp].shape[0]} de plus par rapport au total en {dec_temp}", help= "Cliquez sur le bouton 'Nombre total de questionnaires' dans les paramètres pour personnaliser le delta")
        aut_var=np.round(df["Note de l'accueil"].mean(),2)
        if acc_selected=='Modifier le décalage temporel':
            va=np.round(df["Note de l'accueil"].mean()-df[df["Mois*"]<=dec_temp1]["Note de l'accueil"].mean(),2)
            col2.metric(f"Note moyenne de l'accueil en {year}", f"{aut_var} / 5", f"{va} par rapport au mois de {dec_temp1}")
        elif acc_selected=='Modifier la norme':
            cal=np.round(df["Note de l'accueil"].mean()-new_norm,2)
            col2.metric(f"Note moyenne de l'accueil en {year}", f"{aut_var} / 5", f"{cal} (pour une norme de {new_norm})")
        else:
            cal2=np.round(df["Note de l'accueil"].mean()-4)
            col2.metric(f"Note moyenne de l'accueil en {year}", f"{aut_var} / 5", f"{cal2} (pour une norme de 4)")
        aut_var2=np.round(df["Note de la prise en charge"].mean(),2)
        if pec_selected=='Modifier le décalage temporel ':
            va2=np.round(df[f"Note de la prise en charge en {year}"].mean()-df[df["Mois*"]<=dec_temp2]["Note de la prise en charge"].mean(),2)
            col3.metric(f"Note moyenne de la prise en charge en {year}", f"{aut_var2} / 5", f"{va2} par rapport au mois de {dec_temp2}")
        elif pec_selected=='Modifier la norme ':
            cal3=np.round(df["Note de la prise en charge"].mean()-new_norm2,2)
            col3.metric(f"Note moyenne de la prise en charge en {year}", f"{aut_var2} / 5", f"{cal3} (pour une norme de {new_norm2})")
        else:
            cal4=np.round(df["Note de la prise en charge"].mean()-4,2)
            col3.metric(f"Note moyenne de la prise en charge en {year}", f"{aut_var2} / 5",f"{cal4} (pour une norme de 4)" )

        style_metric_cards(background_color='#0c0c0c',border_left_color="#f7a900",box_shadow=True)

    def palmarès(feat_fil,s,l, critère):
        result_mois = df[((df["Mois de l'année"] >= s) & (df["Mois de l'année"] <= l))].groupby([feat_fil])['Mois'].count().reset_index(name="Total sur le mois")
        result_zone = df.groupby([feat_fil])[feat_fil].count().reset_index(name="Total sur l'année")
        result_mean_acc=df[((df["Mois de l'année"] >= s) & (df["Mois de l'année"] <= l))].groupby([feat_fil])["Note de l'accueil"].mean().reset_index(name="Moy. de l'accueil")
        result_mean_pec=df[((df["Mois de l'année"] >= s) & (df["Mois de l'année"] <= l))].groupby([feat_fil])['Note de la prise en charge'].mean().reset_index(name="Moy. de la prise en charge")
        # Fusion des résultats dans un DataFrame
        if critère=="Total des questionnaires sur l'année":
            d=pd.merge(result_mois, result_zone, on=feat_fil, how='right') 
            d.sort_values(by="Total sur l'année",inplace=True,ascending=False)

        elif critère=="Total des questionnaires sur le mois":
            d=pd.merge(result_mois, result_zone, on=feat_fil, how='right') 
            d.sort_values(by='Total sur le mois',inplace=True,ascending=False)

        elif critère=="Moy. de l'accueil":
            d=pd.merge(result_mois, result_mean_acc, on=feat_fil, how='right')
            d.sort_values(by=critère,inplace=True,ascending=False)
        else:
            d=pd.merge(result_mois, result_mean_pec, on=feat_fil, how='right')
            d.sort_values(by=critère,inplace=True,ascending=False)
        
        return d

    def format_ranking_index(df, index_col='Position'):
        df[index_col] = df.index + 1
        df[index_col] = df[index_col].apply(lambda x: f"{x}ème" if x > 1 else f"{x}er")
        return df.set_index(index_col)
    st.write(" ")
    st.title("KPI mensuels")
    df=df_transitoire
    default=df["Mois de l'année"].unique().sort_values(ascending=False)[0]
    start_val, last_val=st.select_slider("Sélectionner l'intervalle d'analyse", options=df["Mois de l'année"].unique().sort_values(ascending=True), value=[default, default])
    st.header((lambda x: f"Analyse des performances en {start_val}" if start_val == last_val else f"Analyse des performances entre {start_val} et {last_val}")(None),divider='rainbow')
    actu, palm,top=st.columns(3)
    with actu:
        st.subheader("***KPI***", divider='rainbow')
        vt=df[df["Mois de l'année"]==last_val].shape[0]-df[df['Mois']==start_val].shape[0]
        st.metric("Nbre total de questionnaires",df[(df["Mois de l'année"] >= start_val) & (df["Mois de l'année"] <= last_val)].shape[0],f"{vt} quest. de plus entre {start_val} et {last_val}")
        def avis (s,l):
            return int((df[(df["Mois de l'année"] >= s) & (df["Mois de l'année"] <= l)]['Motifs de la note de l\'accueil'].count()+df[(df["Mois de l'année"] >= s) & (df["Mois de l'année"] <= l)]['Motifs de la note de la prise en charge'].count())/2) 
        suggestions = df[df["Mois de l'année"].isin([start_val, last_val])]['Suggestions'].count()
        st.metric("Nbre total d'avis",avis(start_val, last_val),f"{avis(last_val, last_val)-avis(start_val, start_val)} avis de plus entre {start_val} et {last_val}")
        vt2= int(df[(df["Mois de l'année"] >= start_val) & (df["Mois de l'année"] <= last_val)]['Suggestions'].count()-df[(df["Mois de l'année"] >= start_val) & (df["Mois de l'année"] <= last_val)]['Suggestions'].count())
        st.metric("Nbre total de suggestions", suggestions,f"{vt2} sugg. de plus entre {start_val} et {last_val}")
    mean_accueil=df["Note de l'accueil"].mean()
    #df["Nombre de questionnaires remplis par jour"]=mean_accueil[df["Jour"]].values
    with palm:
        st.subheader("***Classement par Zone***",divider="rainbow")
        critère=st.radio("Choisir le critère de classement", ["Total des questionnaires sur l'année", "Total des questionnaires sur le mois", "Moy. de l'accueil", "Moy. de la prise en charge"])
        pa=palmarès('Zone',start_val, last_val, critère)
        pa.reset_index(inplace=True, drop=True)
        st.dataframe(format_ranking_index(pa))
    with top:
        st.subheader("***Classement par agence***",divider='rainbow')
        critère1=st.radio("Choisir le critère de classement ", ["Total des questionnaires sur l'année", "Total des questionnaires sur le mois", "Moy. de l'accueil", "Moy. de la prise en charge"])
        tops=palmarès('Agence',start_val, last_val, critère1)
        tops.reset_index(inplace=True, drop=True)
        st.dataframe(format_ranking_index(tops))

    st.header("Base de données personnalisée",divider="rainbow" )
    df_perso=filter_dataframe(df)
    st.dataframe(df_perso)
    st.sidebar.write("")
    st.sidebar.write("")
    st.sidebar.subheader("PARAMETRES DE LA BASE DE DONNEES")
    df_selected=st.sidebar.radio("***:grey[Choisissez la base de données sur laquel vous souhaitez réaliser les graphiques]***",['Base de données locale', 'Base de données télechargée', 'Base de données personnalisée'])
    if df_selected=='Base de données téléchargée' and uploaded_file is not None:
        df=transf_df(df_uploaded)
    elif df_selected=='Base de données personnalisée':
        df=transf_df(df_perso)
    else:
        df=df[df['Année']==year]

    
    # Histogramme et Camembert sur la même ligne
    cam, hist = st.columns(2,gap='medium')

    with cam:
        st.subheader("CAMEMBERT")
        selected_categorical_variable_p = st.selectbox("***Sélectionnez une variable catégorielle pour le camembert***", ['Agence', "Point de contact","Note de l'accueil","Note de la prise en charge",'Jour','Mois','Zone'], index=1)
        category_counts = df[selected_categorical_variable_p].value_counts()
        fig_pie = px.pie(names=category_counts.index, values=category_counts.values, title=f"Répartition de la variable {selected_categorical_variable_p}",color_discrete_sequence=colors)
        fig_pie.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.25)
        st.plotly_chart(fig_pie, use_container_width=True)

    with hist:
        st.subheader("HISTOGRAMME")
        selected_categorical_variable = st.selectbox("***Sélectionnez la variable catégorielle pour l'histogramme***", ['Agence', "Point de contact","Note de l'accueil", "Note de la prise en charge",'Jour','Mois','Zone'], index=6)
        fig_histogram = px.histogram(df, x=df[selected_categorical_variable], color=df[selected_categorical_variable],title=f"Histogramme de {selected_categorical_variable}",color_discrete_sequence=colors)
        fig_histogram.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.35)
        fig_histogram.update_traces( textfont_color='rgba(255, 255, 255, 1)')
        if selected_categorical_variable=="Mois":
            fig_histogram.update_xaxes(categoryorder='array', categoryarray=order_of_months)
        elif selected_categorical_variable=="Jour":
            fig_histogram.update_xaxes(categoryorder='array', categoryarray=order_of_days)
        fig_histogram.update_xaxes(showticklabels=False)
        st.plotly_chart(fig_histogram,use_container_width=True)




    # Section des analyses croisées




    def barmode_selected(t):
        if t =='empilé':
            a='relative'  
        else: 
            a='group'
        return a

    quant,qual=st.columns(2,gap='medium')


    with quant:
        st.subheader("ANALYSE CROISEE ENTRE VARIABLES NUMERIQUES")
        int_columns = df.select_dtypes(include="int").columns
        float_columns = df.select_dtypes(include="float").columns
        selected_variable_3 = st.selectbox("***Variable 1***", int_columns.union(float_columns))
        selected_variable_4 = st.selectbox("***Variable 2***",int_columns.union(float_columns),index=2)
        occurrences=df.groupby([selected_variable_3, selected_variable_4]).size().reset_index(name='count')
        occurrences['moyenne des deux notes'] = (occurrences[selected_variable_3]+ occurrences[selected_variable_4])/2
        fig_scatter_matrix = px.scatter(occurrences, x=selected_variable_3, y=selected_variable_4, size='count',size_max=70, color='moyenne des deux notes',color_continuous_scale=['red', 'yellow', 'green'])
        fig_scatter_matrix.update_layout(title=f'Nuage de points entre {selected_variable_3} et {selected_variable_4}')
        fig_scatter_matrix.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.15)
        st.plotly_chart(fig_scatter_matrix, use_container_width=True)
    with qual:
        st.subheader("ANALYSE CROISEE ENTRE VARIABLES CATEGORIELLES")
        selected_variable_1 = st.selectbox("***Variable 1***", ['Agence', 'Point de contact','Jour','Mois','Zone'], index=4)
        selected_variable_2 = st.selectbox("***Variable 2***", ['Agence', 'Point de contact',"Note de la prise en charge","Note de l'accueil",'Jour','Mois','Zone'],index=1)
        st.sidebar.write(" ")
        st.sidebar.write(" ")
        st.sidebar.subheader("PARAMETRES DES GRAPHIQUES")
        type_graph=st.sidebar.radio("***:grey[Choisissez le type d'histogramme croisé]***", ['empilé','étalé'])
        if selected_variable_2 in ["Note de l'accueil","Note de la prise en charge"]:
            fig_croisé = px.bar(df.groupby(selected_variable_1)[selected_variable_2].mean().reset_index(), x=selected_variable_1,y=selected_variable_2, color=selected_variable_2,barmode=barmode_selected(type_graph),color_continuous_scale=['red', 'yellow', 'green'],range_color=[0, 5])
        else:
            fig_croisé = px.bar(df, x=selected_variable_1, color=selected_variable_2,barmode=barmode_selected(type_graph),color_discrete_sequence= colors)
            
            if selected_variable_1=="Mois" or selected_variable_2=="Mois":
                fig_croisé.update_xaxes(categoryorder='array', categoryarray=order_of_months)
            elif selected_variable_1=="Jour" or selected_variable_2=="Jour":
                fig_croisé.update_xaxes(categoryorder='array', categoryarray=order_of_days)
        fig_croisé.update_layout(title=f'Graphique en barres groupées - {selected_variable_1 } vs {selected_variable_2 }')
        fig_croisé.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.20)

        st.plotly_chart(fig_croisé,use_container_width=True)


    occurences_day=df["Jour"].value_counts()
    df["Nombre de questionnaires remplis par jour"]=occurences_day[df["Jour"]].values
    occurences_mo=df["Mois"].value_counts()
    df["Nombre de questionnaires remplis dans le mois"]=occurences_mo[df["Mois"]].values
    fig_ann = px.area(df, x="Mois*", y="Nombre de questionnaires remplis dans le mois", color="Jour*",line_group="Nombre de questionnaires remplis par jour",color_discrete_sequence= colors,custom_data=[df["Mois"],df["Jour"],df['Nombre de questionnaires remplis par jour'],df['Nombre de questionnaires remplis dans le mois']])
    fig_ann.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},height= 500,width= 1420)
    fig_ann.update_xaxes(categoryorder='array', categoryarray=order_of_months)
    fig_ann.update_traces(hovertemplate="<b>Mois</b>: %{customdata[0]}<br>"
                                        '<b>Jour</b>: %{customdata[1]}<br>'
                                        '<b>Nbre de quest. enregistrés sur le mois</b>: %{customdata[3]}<br>'
                                        "<b>Nbre de quest. enregistrés ce jour</b>: %{customdata[2]}<br>",
                                        hoverlabel=dict(font=dict(size=16, color='white'))),

    nltk.download('punkt')

    def no_stop_word(string, stopWords):

        """
        Supprime les stop words d'un texte.

        Paramètres
        ----------

        string : chaine de caractère.

        stopWords : liste de mots à exclure.
        """
        string_split=string.lower().split(" ")
        str1=[mot for mot in string_split if mot not in stopWords]
        string =" ".join(str1)
        return string
    stopwords_fr = [
        "a","au", "aux", "avec", "ce", "ces", "dans", "de", "des", "du", "elle", "en", "et", "eux", "il", "je",
        "la", "le", "leur", "lui", "ma", "mais", "me", "même", "mes", "moi", "mon", "ne", "nos", "notre", "nous",
        "on", "ou", "par", "pas", "pour", "qu", "que", "qui", "sa", "se", "ses", "son", "sur", "ta", "te", "tes",
        "toi", "ton", "tu", "un", "une", "vos", "votre", "vous", "c", "d", "j", "l", "à", "m", "n", "s", "t", "y",
        "été", "étée", "étées", "étés", "étant", "étante", "étants", "étantes", "suis", "es", "est", "sommes",
        "êtes", "sont", "serai", "seras", "sera", "serons", "serez", "seront", "serais", "serait", "serions",
        "seriez", "seraient", "étais", "était", "étions", "étiez", "étaient", "fus", "fut", "fûmes", "fûtes",
        "furent", "sois", "soit", "soyons", "soyez", "soient", "fusse", "fusses", "fût", "fussions", "fussiez",
        "fussent", "ayant", "ayante", "ayantes", "ayants", "eu", "eue", "eues", "eus", "ai", "as", "avons", "avez",
        "ont", "aurai", "auras", "aura", "aurons", "aurez", "auront", "aurais", "aurait", "aurions", "auriez",
        "auraient", "avais", "avait", "avions", "aviez", "avaient", "eut", "eûmes", "eûtes", "eurent", "aie", "aies",
        "ait", "nan", "ras","ayons",'rien','none', "ayez", "aient", "eusse", "eusses", "eût", 'non', "eussions","les", "eussiez", "eussent"
    ]

    def stem_cleaner(pandasSeries, stopWords):

        print("#### Nettoyage en cours ####") # Mettre des print vous permet de comprendre où votre code rencontre des problèmes en cas de bug

        # confirmation que chaque article est bien de type str
        pandasSeries = pandasSeries.apply(lambda x : str(x))

        ### COMMENCEZ A CODER ICI! remplacer les 'None' par votre code ###

        # Passage en minuscule
        print("... Passage en minuscule")
        pandasSeries = pandasSeries.apply(lambda x : x.lower())

        # Suppression des accents
        print("... Suppression des accents")
        pandasSeries = pandasSeries.apply(lambda x : unidecode(x))

        # Suppression des caractères spéciaux et numériques
        print("... Suppression des caractères spéciaux et numériques")
        pandasSeries = pandasSeries.apply(lambda x :re.sub(r"[^a-z]+", ' ', x))

        # Suppression des stop words
        print("... Suppression des stop words")
        pandasSeries = pandasSeries.apply(lambda x : no_stop_word(x, stopWords))
        return pandasSeries

    df['comments'] = stem_cleaner(df['Suggestions'], stopwords_fr)
    # Joindre tous les commentaires en une seule chaîne de texte
    def bigramm(df):
        all_comments = ' '.join(df.comments)
        words = word_tokenize(all_comments)
        # Création des bigrammes
        return list(ngrams(words, 2))

    # Compter la fréquence de chaque bigramme
    bigram_counts = Counter(bigramm(df))

    # Sélectionner les 20 bigrammes les plus fréquents pour le treemap
    top_bigrams = dict(sorted(bigram_counts.items(), key=lambda item: item[1], reverse=True)[:20])

    # Créer les données pour le treemap avec les labels et les valeurs de fréquence
    labels = [' '.join(bigram) for bigram in top_bigrams.keys()]
    sizes = list(top_bigrams.values())
    text = [f'Fréquence : {freq}' for freq in sizes]

    


    # Créer le treemap avec Plotly
    fig = go.Figure(go.Treemap(
        labels=labels,
        parents=[""] * len(labels),
        values=sizes,
        text=text,
        hoverinfo='label+text+value'
    ))

    fig.update_layout(
        title='Treemap - Bigrammes les plus fréquents dans les commentaires'
    )
    fig.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},height= 500,width= 1000)



    data = {'Bigrammes': [' '.join(bigram) for bigram in top_bigrams.keys()],
            'Fréquence': list(top_bigrams.values())}

    df3 = pd.DataFrame(data)

    # Créer le graphique à barres horizontales
    fig_bigram = px.bar(df3, x='Fréquence', y='Bigrammes', orientation='h', 
                        title="Histogramme des Bigrammes", 
                        labels={"Fréquence": "Fréquence", "Bigrammes": "Bigrammes"})
    fig_bigram.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},height= 500,width= 1000)

    
    #Wordcloud

    @st.cache_resource
    def word_cloud(df):
        fig_wc, ax = plt.subplots(figsize = (12, 10))

        text = df["comments"].values

        wordcloud_ = WordCloud(background_color=None,
                                collocations=True, # Inclu les bigrammes -> Deux mots cotes à cotes, même couleur même taille
                                width=1200,
                                height=850,
                                mode="RGBA",
                                stopwords=(stopwords_fr)).generate(" ".join(text))

        plt.axis('off')
        plt.title('Nuage de mots des suggestions', fontsize=30)
        ax.imshow(wordcloud_)
        return st.pyplot(fig_wc)
    

    wc,hist_bigramm=st.columns(2,gap='medium')


    with wc:
        but=st.sidebar.button("Nuage de mots de la base filtrée")
        if but:
            df_fil=transf_df(df_perso)
            df_fil["comments"]=stem_cleaner(df_fil['Suggestions'], stopwords_fr)
            word_cloud(df_fil)
        else:
            word_cloud(df)

    with hist_bigramm:
        st.plotly_chart(fig_bigram,use_container_width=True)

    st.plotly_chart(fig,use_container_width=True)
